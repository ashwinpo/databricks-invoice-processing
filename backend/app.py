"""
Invoice Processing API — Databricks App backend.

Endpoints handle the full invoice lifecycle:
  Upload PDF -> ai_parse_document (layout) -> ai_query (field extraction) -> human review -> confirm to Delta

Configuration:
  - app.yaml: env vars for warehouse, volume, table paths, model endpoint
  - invoice_fields.yaml: defines what fields to extract (name, label, description, type)
    The 'description' field is injected into the LLM prompt — more specific = better accuracy.

Key dependencies:
  - Databricks SDK (WorkspaceClient) for file operations and SQL execution
  - ai_parse_document: extracts layout elements with bounding boxes from PDFs
  - ai_query: sends document text + field prompts to a Foundation Model endpoint
"""

from fastapi import FastAPI, HTTPException, UploadFile, File as FastAPIFile, Query
from fastapi.staticfiles import StaticFiles
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, Response
from pydantic import BaseModel
from databricks.sdk import WorkspaceClient
from databricks.sdk.service.sql import StatementState
import os
import yaml
import json
from dotenv import load_dotenv
from typing import List, Optional
import tempfile
import shutil
import base64
import time
from PIL import Image, ImageDraw, ImageFont
import io
import re
import csv

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

def load_yaml_config():
    try:
        with open('app.yaml', 'r') as file:
            config = yaml.safe_load(file)
            yaml_config = {}
            if 'env' in config:
                for env_var in config['env']:
                    yaml_config[env_var['name']] = env_var['value']
            return yaml_config
    except Exception as e:
        print(f"Warning: Could not load app.yaml config: {e}")
        return {}


def load_invoice_fields():
    try:
        with open('invoice_fields.yaml', 'r') as f:
            config = yaml.safe_load(f)
            return config.get('fields', [])
    except Exception as e:
        print(f"Warning: Could not load invoice_fields.yaml: {e}")
        return []


def load_rate_limits():
    defaults = {"max_documents_per_day": 50, "max_files_per_upload": 10}
    try:
        with open('rate_limits.yaml', 'r') as f:
            config = yaml.safe_load(f) or {}
            return {**defaults, **{k: v for k, v in config.items() if isinstance(v, int)}}
    except Exception as e:
        print(f"Warning: Could not load rate_limits.yaml, using defaults: {e}")
        return defaults


YAML_CONFIG = load_yaml_config()
INVOICE_FIELDS = load_invoice_fields()
RATE_LIMITS = load_rate_limits()

load_dotenv()

app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ---------------------------------------------------------------------------
# Pydantic models
# ---------------------------------------------------------------------------

class WriteToTableRequest(BaseModel):
    file_paths: List[str]
    limit: int = 10
    operation_mode: str = 'append'

class QueryDeltaTableRequest(BaseModel):
    file_paths: List[str] = []
    limit: int = 10
    page_number: Optional[int] = None

class VisualizePageRequest(BaseModel):
    file_path: Optional[str] = None
    page_number: Optional[int] = None

class PageMetadataRequest(BaseModel):
    file_paths: List[str] = []

class WarehouseConfigRequest(BaseModel):
    warehouse_id: str

class VolumePathConfigRequest(BaseModel):
    volume_path: str

class DeltaTablePathConfigRequest(BaseModel):
    delta_table_path: str

class ExtractFieldsRequest(BaseModel):
    file_path: str

class BatchExtractFieldsRequest(BaseModel):
    file_paths: List[str]

class ConfirmInvoiceRequest(BaseModel):
    file_path: str
    fields: dict

class DismissErrorRequest(BaseModel):
    file_path: str
    stage: str

class RetryInvoiceRequest(BaseModel):
    file_path: str

# ---------------------------------------------------------------------------
# Databricks client init
# ---------------------------------------------------------------------------

try:
    w = WorkspaceClient()
    warehouse_id = os.getenv("DATABRICKS_WAREHOUSE_ID", YAML_CONFIG.get("DATABRICKS_WAREHOUSE_ID"))
    print(f"Databricks client initialized with warehouse: {warehouse_id}")
except Exception as e:
    print(f"Databricks client initialization failed: {e}")
    w = None
    warehouse_id = None

current_warehouse_id = warehouse_id
current_volume_path = os.getenv("DATABRICKS_VOLUME_PATH", YAML_CONFIG.get("DATABRICKS_VOLUME_PATH"))
current_delta_table_path = os.getenv("DATABRICKS_DELTA_TABLE_PATH", YAML_CONFIG.get("DATABRICKS_DELTA_TABLE_PATH"))
invoice_results_table = os.getenv("INVOICE_RESULTS_TABLE", YAML_CONFIG.get("INVOICE_RESULTS_TABLE"))
ai_query_model = os.getenv("AI_QUERY_MODEL", YAML_CONFIG.get("AI_QUERY_MODEL", "databricks-claude-sonnet-4-20250514"))


def get_uc_volume_path() -> str:
    return current_volume_path or "/Volumes/main/default/ai_functions_demo"

def get_delta_table_path() -> str:
    return current_delta_table_path or "main.default.ai_functions_demo_documents"

def get_invoice_results_table() -> str:
    return invoice_results_table or "main.ai_parse_document_demo.ashwin_invoice_results"

def get_processing_errors_table() -> str:
    """Derive errors table name from the invoice results table."""
    base = get_invoice_results_table()
    return base.rsplit('.', 1)[0] + '.processing_errors'

# ---------------------------------------------------------------------------
# Helper: execute SQL and wait
# ---------------------------------------------------------------------------

def execute_sql(statement: str, wait_timeout: str = '50s', retries: int = 0):
    """Execute SQL and poll until completion. Retries on transient failures."""
    for attempt in range(retries + 1):
        try:
            result = w.statement_execution.execute_statement(
                statement=statement,
                warehouse_id=current_warehouse_id,
                wait_timeout=wait_timeout
            )
            if result.status and result.status.state in [StatementState.PENDING, StatementState.RUNNING]:
                max_wait = 600
                waited = 0
                while result.status.state in [StatementState.PENDING, StatementState.RUNNING] and waited < max_wait:
                    time.sleep(5)
                    waited += 5
                    result = w.statement_execution.get_statement(result.statement_id)
                    print(f"Waiting for SQL completion... ({waited}s) - Status: {result.status.state}")
            # Check for transient failures worth retrying
            if (result.status and result.status.state == StatementState.FAILED
                    and attempt < retries):
                error_str = str(result.status.error) if result.status.error else ""
                if any(t in error_str.lower() for t in ["timeout", "temporarily", "unavailable", "capacity"]):
                    wait_secs = 5 * (attempt + 1)
                    print(f"Transient SQL failure (attempt {attempt + 1}/{retries + 1}), retrying in {wait_secs}s: {error_str[:200]}")
                    time.sleep(wait_secs)
                    continue
            return result
        except Exception as e:
            if attempt < retries:
                wait_secs = 5 * (attempt + 1)
                print(f"SQL exception (attempt {attempt + 1}/{retries + 1}), retrying in {wait_secs}s: {e}")
                time.sleep(wait_secs)
                continue
            raise
    return result


def _log_processing_error(file_path: str, filename: str, stage: str, error_message: str):
    """Log a processing error to the errors Delta table. Best-effort — never raises."""
    try:
        errors_table = get_processing_errors_table()
        execute_sql(f"""
        CREATE TABLE IF NOT EXISTS IDENTIFIER('{errors_table}') (
            file_path STRING,
            filename STRING,
            stage STRING,
            error_message STRING,
            created_at TIMESTAMP,
            resolved BOOLEAN
        ) USING DELTA
        """, '30s')

        fp_esc = file_path.replace("'", "''")
        fn_esc = filename.replace("'", "''")
        err_esc = error_message[:2000].replace("'", "''")
        stage_esc = stage.replace("'", "''")

        execute_sql(f"""
        INSERT INTO IDENTIFIER('{errors_table}')
        (file_path, filename, stage, error_message, created_at, resolved)
        VALUES ('{fp_esc}', '{fn_esc}', '{stage_esc}', '{err_esc}', current_timestamp(), false)
        """, '30s')
        print(f"Logged processing error: {filename} / {stage} / {error_message[:100]}")
    except Exception as log_err:
        print(f"Warning: could not log processing error: {log_err}")


def _get_daily_document_count() -> int:
    """Count documents processed in the last 24 hours from the invoice results table."""
    try:
        results_table = get_invoice_results_table()
        result = execute_sql(f"""
            SELECT COUNT(*) FROM IDENTIFIER('{results_table}')
            WHERE uploaded_at >= current_timestamp() - INTERVAL 24 HOURS
            """, '15s', retries=1)
        if result.result and result.result.data_array:
            return int(result.result.data_array[0][0] or 0)
    except Exception as e:
        # Table may not exist yet — that's fine, count is 0
        if "TABLE_OR_VIEW_NOT_FOUND" not in str(e) and "does not exist" not in str(e).lower():
            print(f"Warning: daily count query failed: {e}")
    return 0


def _check_rate_limits(file_count: int):
    """Raise HTTPException if rate limits would be exceeded. Pass file_count=0 to skip batch check."""
    max_per_upload = RATE_LIMITS.get("max_files_per_upload", 0)
    if max_per_upload > 0 and file_count > max_per_upload:
        raise HTTPException(
            status_code=429,
            detail=f"Upload limited to {max_per_upload} files at a time ({file_count} provided). "
                   f"This limit is configured in rate_limits.yaml."
        )

    max_per_day = RATE_LIMITS.get("max_documents_per_day", 0)
    if max_per_day > 0:
        current_count = _get_daily_document_count()
        remaining = max_per_day - current_count
        if remaining <= 0:
            raise HTTPException(
                status_code=429,
                detail=f"Daily processing limit reached ({max_per_day} documents per 24 hours). "
                       f"Try again later or adjust the limit in rate_limits.yaml."
            )
        if file_count > remaining:
            raise HTTPException(
                status_code=429,
                detail=f"Only {remaining} of {max_per_day} daily documents remaining, "
                       f"but {file_count} files were submitted. "
                       f"Reduce the batch size or try again later."
            )


# ---------------------------------------------------------------------------
# Config endpoints
# ---------------------------------------------------------------------------

@app.get("/api/warehouse-config")
def get_warehouse_config():
    return {
        "warehouse_id": current_warehouse_id,
        "default_warehouse_id": warehouse_id
    }

@app.post("/api/warehouse-config")
def update_warehouse_config(request: WarehouseConfigRequest):
    global current_warehouse_id
    current_warehouse_id = request.warehouse_id
    return {"success": True, "warehouse_id": current_warehouse_id, "message": "Warehouse ID updated"}

@app.get("/api/volume-path-config")
def get_volume_path_config():
    default_path = YAML_CONFIG.get("DATABRICKS_VOLUME_PATH", "/Volumes/main/default/invoices/")
    return {
        "volume_path": current_volume_path or default_path,
        "default_volume_path": default_path
    }

@app.post("/api/volume-path-config")
def update_volume_path_config(request: VolumePathConfigRequest):
    global current_volume_path
    current_volume_path = request.volume_path
    return {"success": True, "volume_path": current_volume_path, "message": "Volume path updated"}

@app.get("/api/delta-table-path-config")
def get_delta_table_path_config():
    default_path = YAML_CONFIG.get("DATABRICKS_DELTA_TABLE_PATH", "main.default.documents")
    return {
        "delta_table_path": current_delta_table_path or default_path,
        "default_delta_table_path": default_path
    }

@app.post("/api/delta-table-path-config")
def update_delta_table_path_config(request: DeltaTablePathConfigRequest):
    global current_delta_table_path
    current_delta_table_path = request.delta_table_path
    return {"success": True, "delta_table_path": current_delta_table_path, "message": "Delta table path updated"}


# ---------------------------------------------------------------------------
# Upload to UC Volume
# ---------------------------------------------------------------------------

@app.post("/api/upload-to-uc")
async def upload_to_uc(files: List[UploadFile] = FastAPIFile(...)):
    if not w:
        raise HTTPException(status_code=500, detail="Databricks connection is not configured.")

    _check_rate_limits(len(files))

    try:
        uploaded_files = []
        base_path = get_uc_volume_path().rstrip('/')

        # Ensure base path and images dir exist
        for dir_path in [base_path, f"{base_path}/images"]:
            try:
                w.files.create_directory(directory_path=dir_path)
            except Exception:
                pass  # Already exists

        # List existing files once — track names across uploads in this batch
        existing_names = set()
        try:
            for entry in w.files.list_directory_contents(directory_path=base_path):
                if entry.name:
                    existing_names.add(entry.name)
        except Exception as e:
            print(f"Could not list directory contents: {e}")

        for file in files:
            with tempfile.NamedTemporaryFile(delete=False) as temp_file:
                shutil.copyfileobj(file.file, temp_file)
                temp_file_path = temp_file.name

            try:
                original_name = file.filename
                stem, suffix = os.path.splitext(original_name)

                final_name = original_name
                counter = 1
                while final_name in existing_names:
                    final_name = f"{stem}_{counter}{suffix}"
                    counter += 1

                uc_file_path = f"{base_path}/{final_name}"
                existing_names.add(final_name)

                with open(temp_file_path, 'rb') as f:
                    w.files.upload(file_path=uc_file_path, contents=f, overwrite=False)

                file_size = os.path.getsize(temp_file_path)
                uploaded_files.append({"name": final_name, "path": uc_file_path, "size": file_size})
            finally:
                os.unlink(temp_file_path)

        return {
            "success": True,
            "uploaded_files": uploaded_files,
            "message": f"Successfully uploaded {len(uploaded_files)} files"
        }
    except Exception as e:
        for file in files:
            _log_processing_error(
                file_path=f"{get_uc_volume_path()}/{file.filename}",
                filename=file.filename or "unknown",
                stage="upload",
                error_message=str(e),
            )
        raise HTTPException(status_code=500, detail=f"Upload failed: {str(e)}")


# ---------------------------------------------------------------------------
# Write to Delta (ai_parse_document)
# ---------------------------------------------------------------------------

@app.post("/api/write-to-delta-table")
def write_to_delta_table(request: WriteToTableRequest):
    if not w:
        raise HTTPException(status_code=500, detail="Databricks connection is not configured.")
    if not current_warehouse_id:
        raise HTTPException(status_code=500, detail="DATABRICKS_WAREHOUSE_ID is not set.")
    if not request.file_paths:
        raise HTTPException(status_code=400, detail="file_paths is required")

    try:
        destination_table = get_delta_table_path()
        print(f"Processing {len(request.file_paths)} file(s) into {destination_table}")

        # Ensure table exists
        create_query = f"""
        CREATE TABLE IF NOT EXISTS IDENTIFIER('{destination_table}') (
            path STRING,
            element_id BIGINT,
            type STRING,
            bbox ARRAY<DOUBLE>,
            page_id STRING,
            content STRING,
            description STRING,
            image_uri STRING
        ) USING DELTA
        """
        execute_sql(create_query, '30s')

        # Delete existing records for these files (append mode dedupe)
        for file_path in request.file_paths:
            dbfs_path = f"dbfs:{file_path}" if file_path.startswith('/Volumes/') else file_path
            execute_sql(f"DELETE FROM IDENTIFIER('{destination_table}') WHERE path = '{dbfs_path}'", '30s')

        # Build file CTE
        dbfs_paths = []
        for fp in request.file_paths:
            dbfs_paths.append(f"dbfs:{fp}" if fp.startswith('/Volumes/') else fp)

        first_file_path = request.file_paths[0]
        base_path = re.sub(r'/[^/]+$', '', first_file_path)

        file_cte = ' UNION ALL '.join([
            f"SELECT path, content FROM READ_FILES('{p}', format => 'binaryFile')"
            for p in dbfs_paths
        ])

        insert_query = f"""
        INSERT INTO IDENTIFIER('{destination_table}')
        WITH file AS (
          {file_cte}
        ),
        parsed as (
          SELECT path,
            ai_parse_document(
              content,
              map('version', '2.0',
                  'imageOutputPath', '{base_path}/images',
                  'descriptionElementTypes', '*')
            ) as parsed
          FROM file
        ),
        pages as (
          SELECT path, id as page_id, cast(image_uri:image_uri as string) as image_uri
          FROM (
            SELECT path, posexplode(try_cast(parsed:document:pages AS ARRAY<VARIANT>)) AS (id, image_uri)
            FROM parsed
            WHERE parsed:document:pages IS NOT NULL AND CAST(parsed:error_status AS STRING) IS NULL
          )
        ),
        elements as (
          SELECT path,
            cast(items:id as int) as element_id,
            cast(items:type as string) as type,
            cast(items:bbox[0]:coord as ARRAY<DOUBLE>) as bbox,
            cast(items:bbox[0]:page_id as int) as page_id,
            CASE WHEN cast(items:type as string) = 'figure' THEN cast(items:description as string)
              ELSE cast(items:content as string) END as content,
            cast(items:description as string) as description
          FROM (
            SELECT path, posexplode(try_cast(parsed:document:elements AS ARRAY<VARIANT>)) AS (idx, items)
            FROM parsed
            WHERE parsed:document:elements IS NOT NULL AND CAST(parsed:error_status AS STRING) IS NULL
          )
        )
        SELECT e.*, p.image_uri FROM elements e
        INNER JOIN pages p ON e.path = p.path AND e.page_id = p.page_id
        """

        print("Executing ai_parse_document INSERT...")
        result = execute_sql(insert_query, retries=2)

        if result.status and result.status.state == StatementState.SUCCEEDED:
            return {
                "success": True,
                "destination_table": destination_table,
                "processed_files": request.file_paths,
                "failed_files": [],
                "message": f"Successfully processed {len(request.file_paths)} files"
            }
        else:
            error_msg = "Processing failed"
            if result.status and result.status.error:
                error_msg += f": {result.status.error}"
            for fp in request.file_paths:
                _log_processing_error(fp, fp.split('/')[-1], "parse", error_msg)
            return {
                "success": False,
                "destination_table": destination_table,
                "processed_files": [],
                "failed_files": [{"file_path": fp, "error": error_msg} for fp in request.file_paths],
                "message": error_msg
            }

    except Exception as e:
        print(f"Delta table write error: {e}")
        for fp in request.file_paths:
            _log_processing_error(fp, fp.split('/')[-1], "parse", str(e))
        raise HTTPException(status_code=500, detail=f"Failed to write to delta table: {str(e)}")


# ---------------------------------------------------------------------------
# Query Delta Table
# ---------------------------------------------------------------------------

@app.post("/api/query-delta-table")
def query_delta_table(request: QueryDeltaTableRequest):
    if not w:
        raise HTTPException(status_code=500, detail="Databricks connection is not configured.")
    if not current_warehouse_id:
        raise HTTPException(status_code=500, detail="DATABRICKS_WAREHOUSE_ID is not set.")

    try:
        destination_table = get_delta_table_path()
        where_conditions = []

        if request.file_paths:
            dbfs_file_paths = [f"dbfs:{fp}" if fp.startswith('/Volumes/') else fp for fp in request.file_paths]
            path_conditions = ", ".join([f"'{fp}'" for fp in dbfs_file_paths])
            where_conditions.append(f"path IN ({path_conditions})")

        if request.page_number is not None:
            where_conditions.append(f"page_id = {request.page_number}")

        where_clause = f"WHERE {' AND '.join(where_conditions)}" if where_conditions else ""

        query = f"""
        SELECT path, element_id, type, cast(bbox as ARRAY<DOUBLE>) as bbox,
               page_id, content, description, image_uri
        FROM IDENTIFIER('{destination_table}')
        {where_clause}
        ORDER BY page_id, element_id
        """

        result = execute_sql(query, '30s', retries=1)

        if result.result and result.result.data_array:
            data = [
                {"path": r[0], "element_id": r[1], "type": r[2], "bbox": r[3],
                 "page_id": r[4], "content": r[5], "description": r[6], "image_uri": r[7]}
                for r in result.result.data_array
            ]
            return {"success": True, "data": data, "total_results": len(data)}
        return {"success": True, "data": [], "message": "No results found"}

    except Exception as e:
        return {"success": False, "data": [], "error": f"Query failed: {str(e)}"}


# ---------------------------------------------------------------------------
# Page Metadata
# ---------------------------------------------------------------------------

@app.post("/api/page-metadata")
def get_page_metadata(request: PageMetadataRequest):
    if not w:
        raise HTTPException(status_code=500, detail="Databricks connection is not configured.")
    if not current_warehouse_id:
        raise HTTPException(status_code=500, detail="DATABRICKS_WAREHOUSE_ID is not set.")

    try:
        destination_table = get_delta_table_path()
        where_conditions = []

        if request.file_paths:
            dbfs_file_paths = [f"dbfs:{fp}" if fp.startswith('/Volumes/') else fp for fp in request.file_paths]
            path_conditions = ", ".join([f"'{fp}'" for fp in dbfs_file_paths])
            where_conditions.append(f"path IN ({path_conditions})")

        where_clause = f"WHERE {' AND '.join(where_conditions)}" if where_conditions else ""

        query = f"""
        SELECT page_id, COUNT(*) as elements_count
        FROM IDENTIFIER('{destination_table}')
        {where_clause}
        GROUP BY page_id ORDER BY page_id
        """

        result = execute_sql(query, '30s', retries=1)

        if result.result and result.result.data_array:
            pages = []
            total_elements = 0
            for row in result.result.data_array:
                page_id = row[0]
                count = int(row[1]) if row[1] else 0
                if page_id is not None:
                    pages.append({"page_id": int(page_id), "page_number": int(page_id) + 1, "elements_count": count})
                    total_elements += count
            return {"success": True, "total_pages": len(pages), "total_elements": total_elements, "pages": pages}
        return {"success": True, "total_pages": 0, "total_elements": 0, "pages": []}

    except Exception as e:
        return {"success": False, "total_pages": 0, "total_elements": 0, "pages": [], "error": str(e)}


# ---------------------------------------------------------------------------
# Visualize Page (bounding boxes)
# ---------------------------------------------------------------------------

@app.post("/api/visualize-page")
def visualize_page(request: VisualizePageRequest):
    if not w:
        raise HTTPException(status_code=500, detail="Databricks connection is not configured.")
    if not current_warehouse_id:
        raise HTTPException(status_code=500, detail="DATABRICKS_WAREHOUSE_ID is not set.")

    try:
        destination_table = get_delta_table_path()
        where_conditions = []

        if request.file_path:
            dbfs_path = f"dbfs:{request.file_path}" if request.file_path.startswith('/Volumes/') else request.file_path
            where_conditions.append(f"path = '{dbfs_path}'")
        if request.page_number is not None:
            where_conditions.append(f"page_id = {request.page_number}")

        where_clause = f"WHERE {' AND '.join(where_conditions)}" if where_conditions else ""

        query = f"""
        SELECT image_uri, page_id,
            collect_list(named_struct(
                'element_id', element_id, 'type', type,
                'bbox', cast(bbox as ARRAY<DOUBLE>),
                'content', content, 'description', description
            )) as element_data_list
        FROM IDENTIFIER('{destination_table}')
        {where_clause}
        GROUP BY image_uri, page_id
        """

        result = execute_sql(query, '30s', retries=1)

        if not result.result or not result.result.data_array:
            return {"success": False, "message": "No elements found"}

        type_color_map = {
            'text': 'blue', 'title': 'red', 'section_header': 'purple',
            'table': 'lime', 'figure': 'magenta', 'page_footer': 'orange', 'page_header': 'orange'
        }

        visualizations = {}
        total_elements = 0

        for row in result.result.data_array:
            image_uri = row[0]
            page_id = row[1]
            element_data_raw = row[2]

            if not image_uri or page_id is None:
                continue

            page_id_str = str(page_id)
            elements = []

            if element_data_raw:
                try:
                    parsed_elements = json.loads(element_data_raw) if isinstance(element_data_raw, str) else element_data_raw
                    for el in parsed_elements:
                        if el and el.get('bbox') and el.get('type'):
                            try:
                                bbox_coords = [float(c) for c in el['bbox']]
                                elements.append({
                                    "element_id": el.get('element_id'),
                                    "type": el['type'],
                                    "bbox": bbox_coords,
                                    "content": el.get('content', ''),
                                    "description": el.get('description', '')
                                })
                            except (ValueError, TypeError):
                                continue
                except (json.JSONDecodeError, TypeError):
                    continue

            if not elements:
                continue

            total_elements += len(elements)

            # Download and annotate image
            try:
                download_path = image_uri[5:] if image_uri.startswith('dbfs:/Volumes/') else image_uri
                image_response = w.files.download(file_path=download_path)

                image_bytes = None
                if hasattr(image_response, 'contents'):
                    if isinstance(image_response.contents, bytes):
                        image_bytes = image_response.contents
                    elif hasattr(image_response.contents, 'read'):
                        image_bytes = image_response.contents.read()
                elif hasattr(image_response, 'content'):
                    image_bytes = image_response.content

                if not image_bytes:
                    continue

                image = Image.open(io.BytesIO(image_bytes))
                image_with_boxes = image.copy()

                for element in elements:
                    bbox = element["bbox"]
                    label = element["type"]
                    color = type_color_map.get(label, 'gray')
                    draw = ImageDraw.Draw(image_with_boxes)
                    draw.rectangle(bbox, outline=color, width=5)

                    try:
                        font = ImageFont.truetype("arial.ttf", 16)
                    except (OSError, IOError):
                        font = ImageFont.load_default()

                    try:
                        bbox_text = draw.textbbox((0, 0), label, font=font)
                        tw, th = bbox_text[2] - bbox_text[0], bbox_text[3] - bbox_text[1]
                    except AttributeError:
                        tw, th = draw.textsize(label, font=font)

                    tx = bbox[0]
                    ty = bbox[1] - th - 2 if bbox[1] - th - 2 > 0 else bbox[1] + 2
                    draw.rectangle([tx - 2, ty - 2, tx + tw + 2, ty + th + 2], fill='white', outline=color)
                    draw.text((tx, ty), label, fill=color, font=font)

                buffer = io.BytesIO()
                image_with_boxes.save(buffer, format='PNG')
                image_base64 = base64.b64encode(buffer.getvalue()).decode('utf-8')

                visualizations[page_id_str] = {
                    "image_base64": image_base64,
                    "elements": elements,
                    "elements_count": len(elements)
                }
            except Exception as e:
                print(f"Error processing page {page_id}: {e}")
                continue

        return {"success": True, "visualizations": visualizations, "total_pages": len(visualizations), "total_elements": total_elements}

    except Exception as e:
        return {"success": False, "message": f"Visualization failed: {str(e)}"}


# ---------------------------------------------------------------------------
# Invoice Fields Config
# ---------------------------------------------------------------------------

@app.get("/api/invoice-fields")
def get_invoice_fields():
    return {"fields": INVOICE_FIELDS}

@app.post("/api/reload-invoice-fields")
def reload_invoice_fields():
    global INVOICE_FIELDS
    INVOICE_FIELDS = load_invoice_fields()
    return {"success": True, "fields": INVOICE_FIELDS, "count": len(INVOICE_FIELDS)}


@app.get("/api/rate-limits")
def get_rate_limits():
    max_per_day = RATE_LIMITS.get("max_documents_per_day", 0)
    current_count = _get_daily_document_count() if max_per_day > 0 else 0
    return {
        "max_documents_per_day": max_per_day,
        "max_files_per_upload": RATE_LIMITS.get("max_files_per_upload", 0),
        "documents_processed_today": current_count,
        "remaining_today": max(0, max_per_day - current_count) if max_per_day > 0 else None,
    }


@app.post("/api/reload-rate-limits")
def reload_rate_limits():
    global RATE_LIMITS
    RATE_LIMITS = load_rate_limits()
    return {"success": True, **RATE_LIMITS}


# ---------------------------------------------------------------------------
# Extract Fields (ai_query)
# ---------------------------------------------------------------------------

@app.post("/api/extract-fields")
def extract_fields(request: ExtractFieldsRequest):
    if not w:
        raise HTTPException(status_code=500, detail="Databricks connection is not configured.")
    if not current_warehouse_id:
        raise HTTPException(status_code=500, detail="DATABRICKS_WAREHOUSE_ID is not set.")

    try:
        destination_table = get_delta_table_path()
        file_path = request.file_path
        dbfs_path = f"dbfs:{file_path}" if file_path.startswith('/Volumes/') else file_path
        dbfs_path_escaped = dbfs_path.replace("'", "''")

        # Build the field descriptions for the prompt — use \n literals for SQL
        field_lines = []
        for f in INVOICE_FIELDS:
            desc = f['description'].replace("'", "''")
            field_lines.append(f"- {f['name']}: {desc}")
        field_descriptions_escaped = "\\n".join(field_lines)

        field_names_escaped = ", ".join([f['name'] for f in INVOICE_FIELDS])

        # Build prompt as a SQL-safe string (single quotes doubled, newlines as \n literals)
        prompt_sql = (
            "You are extracting structured data from an invoice document. "
            "Extract ONLY the following fields from the document text below. "
            "Return a JSON object with exactly these keys. If a value cannot be found, use null.\\n\\n"
            f"Fields to extract:\\n{field_descriptions_escaped}\\n\\n"
            "IMPORTANT:\\n"
            "- Return ONLY a valid JSON object, no other text or markdown.\\n"
            f"- Use the exact field names as keys: {field_names_escaped}\\n"
            "- For dates, use YYYY-MM-DD format when possible.\\n"
            "- For currency amounts, return just the number (e.g. 1234.56 not $1,234.56).\\n\\n"
            "Document text:\\n"
        )

        extract_query = f"""
        WITH doc_content AS (
            SELECT concat_ws('\\n', collect_list(content)) as full_text
            FROM IDENTIFIER('{destination_table}')
            WHERE path = '{dbfs_path_escaped}'
            AND content IS NOT NULL
            AND type IN ('text', 'title', 'section_header', 'table', 'page_header', 'page_footer')
        )
        SELECT ai_query(
            '{ai_query_model}',
            concat('{prompt_sql}', full_text)
        ) as extracted
        FROM doc_content
        WHERE full_text IS NOT NULL
        """

        print(f"Extracting fields for {file_path} using {ai_query_model}")
        result = execute_sql(extract_query, retries=2)

        print(f"Extract result status: {result.status}")

        if result.status and result.status.state == StatementState.SUCCEEDED:
            if result.result and result.result.data_array and len(result.result.data_array) > 0:
                raw_response = result.result.data_array[0][0]
                print(f"Raw ai_query response: {raw_response[:500] if raw_response else 'None'}")

                # Parse the JSON response (handle markdown fences)
                extracted_fields = _parse_ai_response(raw_response)

                if extracted_fields:
                    # Save to invoice results table (non-critical — don't fail the extraction)
                    try:
                        _save_invoice_result(file_path, extracted_fields, status='extracted')
                    except Exception as save_err:
                        print(f"Warning: could not save extraction result: {save_err}")

                    return {
                        "success": True,
                        "file_path": file_path,
                        "fields": extracted_fields
                    }

                err = f"Could not parse AI response: {raw_response[:200] if raw_response else 'empty'}"
                _log_processing_error(file_path, file_path.split('/')[-1], "extract", err)
                return {
                    "success": False,
                    "file_path": file_path,
                    "fields": {f['name']: None for f in INVOICE_FIELDS},
                    "error": err
                }
            else:
                print("Extract query succeeded but no data returned")
                _log_processing_error(file_path, file_path.split('/')[-1], "extract", "No document content found to extract from")
                return {
                    "success": False,
                    "file_path": file_path,
                    "fields": {f['name']: None for f in INVOICE_FIELDS},
                    "error": "No document content found to extract from"
                }

        error_msg = "Extraction query failed"
        if result.status:
            error_msg += f" (state={result.status.state})"
            if result.status.error:
                error_msg += f": {result.status.error}"
        print(f"Extract fields error: {error_msg}")
        _log_processing_error(file_path, file_path.split('/')[-1], "extract", error_msg)
        return {"success": False, "file_path": file_path, "fields": {f['name']: None for f in INVOICE_FIELDS}, "error": error_msg}

    except Exception as e:
        print(f"Field extraction exception: {e}")
        import traceback
        traceback.print_exc()
        _log_processing_error(request.file_path, request.file_path.split('/')[-1], "extract", str(e))
        return {
            "success": False,
            "file_path": request.file_path,
            "fields": {f['name']: None for f in INVOICE_FIELDS},
            "error": str(e)
        }


def _parse_ai_response(raw: str) -> dict:
    """Parse the AI response, handling markdown fences and malformed JSON."""
    if not raw:
        return None
    try:
        # Try direct parse
        return json.loads(raw)
    except json.JSONDecodeError:
        pass
    # Strip markdown fences
    cleaned = re.sub(r'```(?:json)?\s*', '', raw).strip().rstrip('`')
    try:
        return json.loads(cleaned)
    except json.JSONDecodeError:
        pass
    # Try to extract JSON object
    match = re.search(r'\{[^{}]*(?:\{[^{}]*\}[^{}]*)*\}', raw, re.DOTALL)
    if match:
        try:
            return json.loads(match.group())
        except json.JSONDecodeError:
            pass
    return None


@app.post("/api/batch-extract-fields")
def batch_extract_fields(request: BatchExtractFieldsRequest):
    """Extract invoice fields from multiple documents in a single ai_query call.

    Uses failOnError => false so one bad document doesn't kill the batch.
    Returns per-file results with successes and failures separated.
    """
    if not w:
        raise HTTPException(status_code=500, detail="Databricks connection is not configured.")
    if not current_warehouse_id:
        raise HTTPException(status_code=500, detail="DATABRICKS_WAREHOUSE_ID is not set.")
    if not request.file_paths:
        return {"success": True, "results": [], "errors": []}

    try:
        destination_table = get_delta_table_path()

        # Build dbfs paths
        dbfs_paths = []
        path_map = {}  # dbfs_path -> original file_path
        for fp in request.file_paths:
            dbfs = f"dbfs:{fp}" if fp.startswith('/Volumes/') else fp
            dbfs_paths.append(dbfs)
            path_map[dbfs] = fp

        path_filter = ", ".join([f"'{p.replace(chr(39), chr(39)+chr(39))}'" for p in dbfs_paths])

        # Build prompt (same as single-file version)
        field_lines = []
        for f in INVOICE_FIELDS:
            desc = f['description'].replace("'", "''")
            field_lines.append(f"- {f['name']}: {desc}")
        field_descriptions_escaped = "\\n".join(field_lines)
        field_names_escaped = ", ".join([f['name'] for f in INVOICE_FIELDS])

        prompt_sql = (
            "You are extracting structured data from an invoice document. "
            "Extract ONLY the following fields from the document text below. "
            "Return a JSON object with exactly these keys. If a value cannot be found, use null.\\n\\n"
            f"Fields to extract:\\n{field_descriptions_escaped}\\n\\n"
            "IMPORTANT:\\n"
            "- Return ONLY a valid JSON object, no other text or markdown.\\n"
            f"- Use the exact field names as keys: {field_names_escaped}\\n"
            "- For dates, use YYYY-MM-DD format when possible.\\n"
            "- For currency amounts, return just the number (e.g. 1234.56 not $1,234.56).\\n\\n"
            "Document text:\\n"
        )

        batch_query = f"""
        WITH doc_content AS (
            SELECT path, concat_ws('\\n', collect_list(content)) as full_text
            FROM IDENTIFIER('{destination_table}')
            WHERE path IN ({path_filter})
            AND content IS NOT NULL
            AND type IN ('text', 'title', 'section_header', 'table', 'page_header', 'page_footer')
            GROUP BY path
        )
        SELECT
            path,
            ai_query(
                '{ai_query_model}',
                concat('{prompt_sql}', full_text),
                failOnError => false
            ) as ai_response
        FROM doc_content
        WHERE full_text IS NOT NULL
        """

        print(f"Batch extracting fields for {len(request.file_paths)} files using {ai_query_model}")
        result = execute_sql(batch_query, retries=2)

        if not (result.status and result.status.state == StatementState.SUCCEEDED):
            error_msg = "Batch extraction query failed"
            if result.status and result.status.error:
                error_msg += f": {result.status.error}"
            for fp in request.file_paths:
                _log_processing_error(fp, fp.split('/')[-1], "extract", error_msg)
            return {"success": False, "results": [], "errors": [
                {"file_path": fp, "error": error_msg} for fp in request.file_paths
            ]}

        # Parse results — each row is (path, ai_response_struct)
        # ai_response with failOnError=false returns a STRUCT rendered as JSON: {"response": "...", "error": "..."}
        successes = []
        errors = []
        seen_paths = set()

        if result.result and result.result.data_array:
            for row in result.result.data_array:
                dbfs_path = row[0]
                original_path = path_map.get(dbfs_path, dbfs_path)
                filename = original_path.split('/')[-1]
                seen_paths.add(dbfs_path)

                raw_ai = row[1]

                # Parse the failOnError struct — comes back as JSON string
                ai_response = None
                ai_error = None
                if raw_ai:
                    try:
                        struct = json.loads(raw_ai) if isinstance(raw_ai, str) else raw_ai
                        ai_response = struct.get("response")
                        ai_error = struct.get("error")
                    except (json.JSONDecodeError, TypeError, AttributeError):
                        # Might be a plain string response (non-struct)
                        ai_response = raw_ai

                if ai_error:
                    _log_processing_error(original_path, filename, "extract", ai_error)
                    errors.append({"file_path": original_path, "error": ai_error})
                    continue

                extracted_fields = _parse_ai_response(ai_response)
                if extracted_fields:
                    try:
                        _save_invoice_result(original_path, extracted_fields, status='extracted')
                    except Exception as save_err:
                        print(f"Warning: could not save extraction result for {filename}: {save_err}")
                    successes.append({"file_path": original_path, "fields": extracted_fields})
                else:
                    err = f"Could not parse AI response: {str(ai_response)[:200] if ai_response else 'empty'}"
                    _log_processing_error(original_path, filename, "extract", err)
                    errors.append({"file_path": original_path, "error": err})

        # Any paths not in result = no content found
        for dbfs_path, original_path in path_map.items():
            if dbfs_path not in seen_paths:
                err = "No document content found to extract from"
                _log_processing_error(original_path, original_path.split('/')[-1], "extract", err)
                errors.append({"file_path": original_path, "error": err})

        return {"success": True, "results": successes, "errors": errors}

    except Exception as e:
        print(f"Batch extraction exception: {e}")
        import traceback
        traceback.print_exc()
        for fp in request.file_paths:
            _log_processing_error(fp, fp.split('/')[-1], "extract", str(e))
        return {"success": False, "results": [], "errors": [
            {"file_path": fp, "error": str(e)} for fp in request.file_paths
        ]}


def _save_invoice_result(file_path: str, fields: dict, status: str = 'extracted'):
    """Save or update an invoice result in the results table.

    Raises on failure so callers can report errors to the user.
    For non-critical callers (e.g. extract-fields), catch externally.
    """
    results_table = get_invoice_results_table()
    filename = file_path.split('/')[-1]
    fields_json = json.dumps(fields).replace("'", "''")
    file_path_escaped = file_path.replace("'", "''")
    filename_escaped = filename.replace("'", "''")

    # Ensure table exists
    execute_sql(f"""
    CREATE TABLE IF NOT EXISTS IDENTIFIER('{results_table}') (
        file_path STRING,
        filename STRING,
        status STRING,
        extracted_fields STRING,
        confirmed_fields STRING,
        uploaded_at TIMESTAMP,
        confirmed_at TIMESTAMP
    ) USING DELTA
    """, '30s')

    # Upsert: delete then insert (simpler than UPDATE for Databricks SQL)
    execute_sql(f"DELETE FROM IDENTIFIER('{results_table}') WHERE file_path = '{file_path_escaped}'", '30s')

    if status == 'confirmed':
        execute_sql(f"""
        INSERT INTO IDENTIFIER('{results_table}')
        (file_path, filename, status, confirmed_fields, confirmed_at)
        VALUES ('{file_path_escaped}', '{filename_escaped}', '{status}', '{fields_json}', current_timestamp())
        """, '30s')
    else:
        execute_sql(f"""
        INSERT INTO IDENTIFIER('{results_table}')
        (file_path, filename, status, extracted_fields, uploaded_at)
        VALUES ('{file_path_escaped}', '{filename_escaped}', '{status}', '{fields_json}', current_timestamp())
        """, '30s')

    print(f"Saved invoice result: {file_path} ({status})")


# ---------------------------------------------------------------------------
# Invoice Queue
# ---------------------------------------------------------------------------

@app.get("/api/invoice-queue")
def get_invoice_queue():
    if not w:
        raise HTTPException(status_code=500, detail="Databricks connection is not configured.")
    if not current_warehouse_id:
        raise HTTPException(status_code=500, detail="DATABRICKS_WAREHOUSE_ID is not set.")

    try:
        results_table = get_invoice_results_table()

        # Check if table exists
        try:
            result = execute_sql(f"""
                SELECT file_path, filename, status, extracted_fields, confirmed_fields, uploaded_at, confirmed_at
                FROM IDENTIFIER('{results_table}')
                ORDER BY uploaded_at DESC
                """, '30s', retries=1)

            invoices = []
            if result.result and result.result.data_array:
                for row in result.result.data_array:
                    extracted = None
                    confirmed = None
                    try:
                        if row[3]:
                            extracted = json.loads(row[3])
                    except (json.JSONDecodeError, TypeError):
                        pass
                    try:
                        if row[4]:
                            confirmed = json.loads(row[4])
                    except (json.JSONDecodeError, TypeError):
                        pass

                    invoices.append({
                        "file_path": row[0],
                        "filename": row[1],
                        "status": row[2],
                        "extracted_fields": extracted,
                        "confirmed_fields": confirmed,
                        "uploaded_at": row[5],
                        "confirmed_at": row[6]
                    })

            return {"success": True, "invoices": invoices}
        except Exception as e:
            if "TABLE_OR_VIEW_NOT_FOUND" in str(e) or "does not exist" in str(e).lower():
                return {"success": True, "invoices": []}
            raise

    except Exception as e:
        print(f"Invoice queue error: {e}")
        return {"success": False, "invoices": [], "error": str(e)}


# ---------------------------------------------------------------------------
# Confirm Invoice
# ---------------------------------------------------------------------------

@app.post("/api/confirm-invoice")
def confirm_invoice(request: ConfirmInvoiceRequest):
    if not w:
        raise HTTPException(status_code=500, detail="Databricks connection is not configured.")

    try:
        _save_invoice_result(request.file_path, request.fields, status='confirmed')
        return {"success": True, "file_path": request.file_path, "message": "Invoice confirmed"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to confirm invoice: {str(e)}")


# ---------------------------------------------------------------------------
# Export Invoices (CSV / Excel)
# ---------------------------------------------------------------------------

@app.get("/api/export-invoices")
def export_invoices(export_format: str = Query("csv", alias="format")):
    if not w:
        raise HTTPException(status_code=500, detail="Databricks connection is not configured.")
    if not current_warehouse_id:
        raise HTTPException(status_code=500, detail="DATABRICKS_WAREHOUSE_ID is not set.")

    if export_format not in ("csv", "xlsx"):
        raise HTTPException(status_code=400, detail="Format must be 'csv' or 'xlsx'")

    try:
        results_table = get_invoice_results_table()
        result = execute_sql(f"""
            SELECT file_path, filename, status, extracted_fields, confirmed_fields,
                   uploaded_at, confirmed_at
            FROM IDENTIFIER('{results_table}')
            ORDER BY confirmed_at DESC NULLS LAST, uploaded_at DESC
            """, '30s', retries=1)

        rows = []
        if result.result and result.result.data_array:
            for row in result.result.data_array:
                # Use confirmed_fields if available, else extracted_fields
                fields = {}
                try:
                    if row[4]:
                        fields = json.loads(row[4])
                    elif row[3]:
                        fields = json.loads(row[3])
                except (json.JSONDecodeError, TypeError):
                    pass
                rows.append({
                    "filename": row[1],
                    "status": row[2],
                    **{f['name']: fields.get(f['name'], '') for f in INVOICE_FIELDS},
                    "uploaded_at": row[5] or '',
                    "confirmed_at": row[6] or '',
                })

        # Build column headers
        columns = ["filename", "status"] + [f['label'] for f in INVOICE_FIELDS] + ["uploaded_at", "confirmed_at"]
        data_keys = ["filename", "status"] + [f['name'] for f in INVOICE_FIELDS] + ["uploaded_at", "confirmed_at"]

        if export_format == "csv":
            output = io.StringIO()
            writer = csv.writer(output)
            writer.writerow(columns)
            for r in rows:
                writer.writerow([r.get(k, '') for k in data_keys])

            return Response(
                content=output.getvalue(),
                media_type="text/csv",
                headers={"Content-Disposition": "attachment; filename=invoices_export.csv"}
            )

        else:  # xlsx
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            except ImportError:
                raise HTTPException(status_code=500, detail="openpyxl not installed — cannot generate Excel")

            wb = Workbook()
            ws = wb.active
            ws.title = "Invoice Export"

            # Header style
            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
            thin_border = Border(
                left=Side(style='thin', color='D0D0D0'),
                right=Side(style='thin', color='D0D0D0'),
                top=Side(style='thin', color='D0D0D0'),
                bottom=Side(style='thin', color='D0D0D0'),
            )

            # Write headers
            for col_idx, header in enumerate(columns, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
                cell.border = thin_border

            # Write data
            status_fills = {
                'confirmed': PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
                'extracted': PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
            }
            for row_idx, r in enumerate(rows, 2):
                for col_idx, key in enumerate(data_keys, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=r.get(key, ''))
                    cell.border = thin_border
                    if key == 'status' and r.get('status') in status_fills:
                        cell.fill = status_fills[r['status']]

            # Auto-fit column widths
            for col_idx, key in enumerate(data_keys, 1):
                max_len = len(columns[col_idx - 1])
                for r in rows:
                    val = str(r.get(key, ''))
                    if len(val) > max_len:
                        max_len = len(val)
                ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 4, 40)

            # Freeze header row
            ws.freeze_panes = 'A2'

            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)

            return Response(
                content=buffer.getvalue(),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": "attachment; filename=invoices_export.xlsx"}
            )

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Export failed: {str(e)}")


# ---------------------------------------------------------------------------
# Processing Errors
# ---------------------------------------------------------------------------

@app.get("/api/processing-errors")
def get_processing_errors():
    """Return unresolved processing errors."""
    if not w:
        raise HTTPException(status_code=500, detail="Databricks connection is not configured.")
    if not current_warehouse_id:
        raise HTTPException(status_code=500, detail="DATABRICKS_WAREHOUSE_ID is not set.")

    try:
        errors_table = get_processing_errors_table()
        try:
            result = execute_sql(f"""
                SELECT file_path, filename, stage, error_message, created_at
                FROM IDENTIFIER('{errors_table}')
                WHERE resolved = false
                ORDER BY created_at DESC
                """, '30s', retries=1)

            errors = []
            if result.result and result.result.data_array:
                for row in result.result.data_array:
                    errors.append({
                        "file_path": row[0],
                        "filename": row[1],
                        "stage": row[2],
                        "error_message": row[3],
                        "created_at": row[4],
                    })
            return {"success": True, "errors": errors}
        except Exception as e:
            if "TABLE_OR_VIEW_NOT_FOUND" in str(e) or "does not exist" in str(e).lower():
                return {"success": True, "errors": []}
            raise

    except Exception as e:
        print(f"Processing errors query failed: {e}")
        return {"success": False, "errors": [], "error": str(e)}


@app.post("/api/dismiss-error")
def dismiss_error(request: DismissErrorRequest):
    """Mark errors for a given file_path + stage as resolved."""
    if not w:
        raise HTTPException(status_code=500, detail="Databricks connection is not configured.")

    try:
        errors_table = get_processing_errors_table()
        fp_esc = request.file_path.replace("'", "''")
        stage_esc = request.stage.replace("'", "''")
        execute_sql(f"""
            UPDATE IDENTIFIER('{errors_table}')
            SET resolved = true
            WHERE file_path = '{fp_esc}' AND stage = '{stage_esc}' AND resolved = false
            """, '30s')
        return {"success": True}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to dismiss error: {str(e)}")


@app.post("/api/retry-invoice")
def retry_invoice(request: RetryInvoiceRequest):
    """Re-run parse + extract for a previously failed invoice.

    Resolves existing errors for the file and kicks off processing again.
    """
    if not w:
        raise HTTPException(status_code=500, detail="Databricks connection is not configured.")
    if not current_warehouse_id:
        raise HTTPException(status_code=500, detail="DATABRICKS_WAREHOUSE_ID is not set.")

    file_path = request.file_path

    # Resolve any existing errors for this file
    try:
        errors_table = get_processing_errors_table()
        fp_esc = file_path.replace("'", "''")
        execute_sql(f"""
            UPDATE IDENTIFIER('{errors_table}')
            SET resolved = true
            WHERE file_path = '{fp_esc}' AND resolved = false
            """, '30s')
    except Exception:
        pass  # Table may not exist yet — that's fine

    # Re-run parse
    try:
        parse_req = WriteToTableRequest(file_paths=[file_path])
        parse_result = write_to_delta_table(parse_req)
        if not parse_result.get("success"):
            return {"success": False, "stage": "parse", "error": parse_result.get("message", "Parse failed")}
    except HTTPException as he:
        return {"success": False, "stage": "parse", "error": he.detail}
    except Exception as e:
        return {"success": False, "stage": "parse", "error": str(e)}

    # Re-run extract
    try:
        extract_req = ExtractFieldsRequest(file_path=file_path)
        extract_result = extract_fields(extract_req)
        if not extract_result.get("success"):
            return {"success": False, "stage": "extract", "error": extract_result.get("error", "Extract failed")}
    except Exception as e:
        return {"success": False, "stage": "extract", "error": str(e)}

    return {"success": True, "file_path": file_path, "fields": extract_result.get("fields", {})}


# ---------------------------------------------------------------------------
# Static file serving
# ---------------------------------------------------------------------------

static_dir = os.getenv("STATIC_FILES_PATH", YAML_CONFIG.get("STATIC_FILES_PATH", "static"))
if not os.path.exists(static_dir):
    static_dir = "static"

print(f"Serving static files from: {static_dir}")

try:
    if os.path.exists(f"{static_dir}/_next"):
        app.mount("/_next", StaticFiles(directory=f"{static_dir}/_next"), name="nextjs-assets")
except Exception as e:
    print(f"Failed to mount static files: {e}")


@app.get("/favicon.ico")
def favicon():
    path = f"{static_dir}/favicon.ico"
    if os.path.exists(path):
        return FileResponse(path)
    raise HTTPException(status_code=404, detail="Favicon not found")


@app.get("/{asset_path:path}")
def serve_static_asset(asset_path: str):
    if any(asset_path.endswith(ext) for ext in ['.js', '.css', '.woff2', '.svg', '.png', '.ico', '.jpg', '.jpeg']):
        file_path = f"{static_dir}/{asset_path}"
        if os.path.exists(file_path):
            return FileResponse(file_path)
        raise HTTPException(status_code=404, detail=f"Not found: {asset_path}")

    # Page routes — serve appropriate HTML
    if asset_path.startswith("document-intelligence"):
        for p in [f"{static_dir}/document-intelligence/index.html", f"{static_dir}/document-intelligence.html"]:
            if os.path.exists(p):
                return FileResponse(p)

    index = f"{static_dir}/index.html"
    if os.path.exists(index):
        return FileResponse(index)
    raise HTTPException(status_code=404, detail="Page not found")
